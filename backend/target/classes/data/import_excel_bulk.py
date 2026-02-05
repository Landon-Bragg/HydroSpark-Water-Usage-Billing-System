import openpyxl
import mysql.connector
from datetime import datetime
import uuid

# Database configuration
DB_CONFIG = {
    'host': 'localhost',
    'port': 3307,
    'user': 'Hydrospark',
    'password': 'Hydrospark123',
    'database': 'hydrospark'
}

# UPDATE THIS PATH TO YOUR FULL EXCEL FILE
EXCEL_FILE = r'C:\Users\landon_bragg1\Downloads\Hydrospark-system\backend\src\main\resources\data\DailyUsage.xlsx'

# Batch size for inserts (larger = faster, but uses more memory)
BATCH_SIZE = 5000

def main():
    print("🚀 Starting FAST bulk import for 1M+ rows...")
    print(f"📂 File: {EXCEL_FILE}")
    print(f"⚡ Batch size: {BATCH_SIZE:,} rows\n")
    
    # Connect to MySQL
    print("📊 Connecting to MySQL...")
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    
    # Load Excel file
    print("📂 Loading Excel file (this may take a moment for large files)...")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    sheet = wb['Sheet1']
    
    # Stats
    stats = {
        'total_rows': 0,
        'customers_created': 0,
        'meters_created': 0,
        'readings_inserted': 0,
        'errors': 0
    }
    def normalize_customer_type(raw):
        if raw is None:
            return "RESIDENTIAL"

        s = str(raw).strip().upper()

        # Handle numeric/garbage values like 0
        if s in ("0", "NONE", "N/A", ""):
            return "RESIDENTIAL"

        # Map known Excel values to DB enum values
        if s == "RESIDENTIAL":
            return "RESIDENTIAL"
        if s == "COMMERCIAL":
            return "COMMERCIAL"
        if s == "MUNICIPAL":
            # If your DB enum doesn't have MUNICIPAL, map it to COMMERCIAL (or RESIDENTIAL)
            return "COMMERCIAL"

        # Fallback
        return "RESIDENTIAL"


    # Step 1: Collect unique customers and meters
    print("\n📋 STEP 1: Analyzing unique customers and meters...")
    customers_to_create = {}
    meters_to_create = {}
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        stats['total_rows'] += 1
        
        if stats['total_rows'] % 10000 == 0:
            print(f"   Analyzed {stats['total_rows']:,} rows...")
        
        try:
            customer_name = row[0]
            mailing_address = row[1]
            location_id = str(int(row[2])) if row[2] else None
            customer_type = row[3]
            cycle_number = int(row[4]) if row[4] else 20
            phone = row[5]
            
            if not customer_name or not location_id:
                continue
            
            # Parse name
            name_parts = customer_name.split(' ', 1)
            first_name = name_parts[0] if len(name_parts) > 0 else 'Unknown'
            last_name = name_parts[1] if len(name_parts) > 1 else 'Customer'
            
            # Parse address
            address_parts = mailing_address.split(',') if mailing_address else []
            service_address = address_parts[0].strip() if len(address_parts) > 0 else '123 Main St'
            city_state_zip = address_parts[1].strip() if len(address_parts) > 1 else 'Dallas, TX 75201'
            city_parts = city_state_zip.split(',')
            city = city_parts[0].strip() if len(city_parts) > 0 else 'Dallas'
            state_zip = city_parts[1].strip().split() if len(city_parts) > 1 else ['TX', '75201']
            state = state_zip[0] if len(state_zip) > 0 else 'TX'
            zip_code = state_zip[1] if len(state_zip) > 1 else '75201'
            
            account_number = f"ACCT-{location_id}"
            
            # Store customer data
            if account_number not in customers_to_create:
                customers_to_create[account_number] = {
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': f"{first_name.lower()}.{last_name.lower()}.{location_id}@example.com",
                    'phone': phone,
                    'service_address': service_address,
                    'city': city,
                    'state': state,
                    'zip_code': zip_code,
                    'mailing_address': mailing_address,
                    'cycle_number': cycle_number,
                    'customer_type': normalize_customer_type(customer_type),
                    'location_id': location_id
                }
            
            # Store meter data
            if location_id not in meters_to_create:
                meters_to_create[location_id] = {
                    'account_number': account_number,
                    'service_address': service_address,
                    'city': city,
                    'state': state,
                    'zip_code': zip_code
                }
        
        except Exception as e:
            stats['errors'] += 1
    
    print(f"\n✅ Found {len(customers_to_create)} unique customers")
    print(f"✅ Found {len(meters_to_create)} unique meters")
    
    # Step 2: Bulk insert customers
    print("\n📋 STEP 2: Creating customers...")
    customer_ids = {}
    
    for account_number, cust_data in customers_to_create.items():
        # Check if exists
        cursor.execute("SELECT id FROM customers WHERE account_number = %s", (account_number,))
        result = cursor.fetchone()
        
        if result:
            customer_ids[account_number] = result[0]
        else:
            customer_id = str(uuid.uuid4())
            customer_ids[account_number] = customer_id
            
            cursor.execute("""
                INSERT INTO customers (
                    id, account_number, name, first_name, last_name, email, phone,
                    service_address, city, state, zip_code,
                    mailing_address_line1, mailing_city, mailing_state, mailing_zip,
                    billing_cycle_number, customer_type
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                customer_id, account_number,
                f"{cust_data['first_name']} {cust_data['last_name']}".strip(),
                cust_data['first_name'], cust_data['last_name'],
                cust_data['email'], cust_data['phone'],
                cust_data['service_address'], cust_data['city'],
                cust_data['state'], cust_data['zip_code'],
                cust_data['mailing_address'], cust_data['city'],
                cust_data['state'], cust_data['zip_code'],
                cust_data['cycle_number'], cust_data['customer_type']
            ))

            stats['customers_created'] += 1
    
    conn.commit()
    print(f"✅ Created {stats['customers_created']} new customers")
    
    # Step 3: Bulk insert meters
    print("\n📋 STEP 3: Creating meters...")
    meter_ids = {}
    
    for location_id, meter_data in meters_to_create.items():
        # Check if exists
        cursor.execute("SELECT id FROM meters WHERE external_location_id = %s", (location_id,))
        result = cursor.fetchone()
        
        if result:
            meter_ids[location_id] = result[0]
        else:
            meter_id = str(uuid.uuid4())
            meter_ids[location_id] = meter_id
            customer_id = customer_ids[meter_data['account_number']]
            
            cursor.execute("""
                INSERT INTO meters (
                    id, customer_id, external_location_id,
                    service_address_line1, service_city, service_state, service_zip,
                    status
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                meter_id, customer_id, location_id,
                meter_data['service_address'], meter_data['city'],
                meter_data['state'], meter_data['zip_code'],
                'ACTIVE'
            ))
            stats['meters_created'] += 1
    
    conn.commit()
    print(f"✅ Created {stats['meters_created']} new meters")
    
    # Step 4: Bulk insert readings (FAST!)
    print(f"\n📋 STEP 4: Importing meter readings (batches of {BATCH_SIZE:,})...")
    
    # Reload workbook for second pass
    wb.close()
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    sheet = wb['Sheet1']
    
    reading_batch = []
    row_count = 0
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_count += 1
        
        try:
            location_id = str(int(row[2])) if row[2] else None
            year = int(row[8]) if row[8] else None
            month = int(row[9]) if row[9] else None
            day = int(row[10]) if row[10] else None
            usage_ccf = float(row[11]) if row[11] else 0.0
            
            if not location_id or not year:
                continue
            
            meter_id = meter_ids.get(location_id)
            if not meter_id:
                continue
            
            reading_date = f"{year:04d}-{month:02d}-{day:02d}"
            reading_id = str(uuid.uuid4())
            
            reading_batch.append((
                reading_id, meter_id, reading_date, usage_ccf,
                'EXCEL_IMPORT', datetime.now()
            ))
            
            # Insert batch when full
            if len(reading_batch) >= BATCH_SIZE:
                cursor.executemany("""
                    INSERT IGNORE INTO meter_readings 
                    (id, meter_id, reading_date, usage_ccf, source, created_at)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, reading_batch)
                stats['readings_inserted'] += len(reading_batch)
                conn.commit()
                
                print(f"   ✓ Inserted {stats['readings_inserted']:,} readings ({row_count:,} rows processed)")
                reading_batch = []
        
        except Exception as e:
            stats['errors'] += 1
    
    # Insert remaining readings
    if reading_batch:
        cursor.executemany("""
            INSERT IGNORE INTO meter_readings 
            (id, meter_id, reading_date, usage_ccf, source, created_at)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, reading_batch)
        stats['readings_inserted'] += len(reading_batch)
        conn.commit()
    
    # Close connections
    cursor.close()
    conn.close()
    wb.close()
    
    # Print summary
    print("\n" + "="*60)
    print("✅ BULK IMPORT COMPLETE!")
    print("="*60)
    print(f"Total rows processed:   {stats['total_rows']:,}")
    print(f"Customers created:      {stats['customers_created']:,}")
    print(f"Meters created:         {stats['meters_created']:,}")
    print(f"Readings inserted:      {stats['readings_inserted']:,}")
    print(f"Errors:                 {stats['errors']:,}")
    print("="*60)
    print("\n💡 Run verify_import.py to check the results!")

if __name__ == '__main__':
    main()